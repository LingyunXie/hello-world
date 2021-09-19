import os
import openpyxl
from openpyxl.utils import column_index_from_string


# 递归获取path及其子目录下所有类型为filetype的文件的路径，保存在列表file_path_list中,上述路径都为绝对路径
# eg: path:'D:\\aabc', filetype:'.car'
def listdir(path, filetype, file_path_list):
    for file in os.listdir(path):
        file_path = os.path.join(path, file)
        if os.path.isdir(file_path):
            listdir(file_path, filetype, file_path_list)
        elif os.path.splitext(file_path)[1] == filetype:
            file_path_list.append(file_path)


# 循环car包路径的列表，获取路径里携带的plugin包名和car包名
def save_data_to_excel(file_path_list, worksheet, row_index):
    for dirPath in file_path_list:
        print(dirPath)
        path_list = dirPath.split('\\')
        car_name = path_list[len(path_list) - 1]
        print(car_name)
        plugin_name = path_list[len(path_list) - 3]
        print(plugin_name)
        # 保存 car包名和plugin名到对应列
        worksheet.cell(row_index, pluginNameColumn).value = plugin_name
        worksheet.cell(row_index, carNameColumn).value = car_name
        row_index += 1


path_list = []

# 递归获取 D:\\aabc 及其子目录下所有类型为.car文件的路径，保存在列表path_list中
listdir('D:\\aabc', '.car', path_list)
print(path_list)

# 新建excel，打开第一个sheet页，从第二行
wb = openpyxl.Workbook()
ws = wb.active

# 获取插件包和car包所在列
pluginNameColumn = column_index_from_string('C')
carNameColumn = column_index_from_string('D')

# 填充第一行内容
ws.cell(1, pluginNameColumn).value = 'plugin包'
ws.cell(1, carNameColumn).value = 'car包'

# 从excel第二行开始保存path_list里提取出来的plugin包名称和car包
save_data_to_excel(path_list, ws, 2)


wb.save('文件夹提取结果.xlsx')  # 保存时候另存为的文件名
print('文件夹提取结束')
